import csv
import io
from datetime import datetime, timezone
from typing import Any
from fastapi import APIRouter, Depends, Query, UploadFile, File, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, or_, func
from openpyxl import load_workbook

from app.db.base import get_db
from app.core.deps import get_current_user
from app.db.models.user import User, UserRole
from app.db.models.profile_update_request import ProfileUpdateRequest
from app.schemas.user_schemas import UserCreate, UserResponse, UserUpdate
from app.schemas.profile_update_schemas import (
    ProfileUpdateRequestCreate,
    ProfileUpdateRequestReview,
    ProfileUpdateRequestResponse
)
from app.core.exceptions import ForbiddenError, NotFoundError, ValidationError, ConflictError
from app.core.security import hash_password
from app.services.user_service import UserService

router = APIRouter(tags=["Employees"])


def _require_hr_or_admin(user: User):
    if user.role not in [UserRole.ADMIN, UserRole.HR, UserRole.ORG_ADMIN]:
        raise ForbiddenError("HR or Admin access required.")


# ─── Employee Management Endpoints (Admin/HR Only) ───────────────────────────

@router.get("/employees", response_model=list[UserResponse])
async def list_employees(
    response: Response,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    search: str | None = Query(None),
    department: str | None = Query(None),
    status: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100)
):
    _require_hr_or_admin(user)
    
    q_base = select(User).where(User.tenant_id == user.tenant_id, User.deleted_at.is_(None))
    
    if search:
        search_pattern = f"%{search}%"
        q_base = q_base.where(
            or_(
                User.name.ilike(search_pattern),
                User.email.ilike(search_pattern),
                User.employee_id.ilike(search_pattern)
            )
        )
    
    if department:
        q_base = q_base.where(User.department.ilike(f"%{department}%"))
        
    if status:
        q_base = q_base.where(User.status == status)
        
    # Get total count
    count_q = select(func.count()).select_from(q_base.subquery())
    count_result = await db.execute(count_q)
    total = count_result.scalar_one()
    
    q = q_base.order_by(User.created_at.desc())
    q = q.offset((page - 1) * page_size).limit(page_size)
    
    result = await db.execute(q)
    employees = result.scalars().all()
    
    response.headers["X-Total-Count"] = str(total)
    response.headers["Access-Control-Expose-Headers"] = "X-Total-Count"
    return [UserResponse.model_validate(emp) for emp in employees]


@router.post("/employees", response_model=UserResponse, status_code=201)
async def create_employee(
    payload: UserCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user)
):
    _require_hr_or_admin(user)
    return await UserService.create_user(db, user.tenant_id, payload)


@router.get("/employees/export")
async def export_employees(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    department: str | None = Query(None),
    status: str | None = Query(None)
):
    _require_hr_or_admin(user)
    
    q = select(User).where(User.tenant_id == user.tenant_id, User.deleted_at.is_(None))
    if department:
        q = q.where(User.department.ilike(f"%{department}%"))
    if status:
        q = q.where(User.status == status)
    
    result = await db.execute(q)
    employees = result.scalars().all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Employee ID", "Name", "Email", "Role", "Department", 
        "Designation", "Phone", "Date of Joining", "Status"
    ])
    
    for emp in employees:
        doj_str = emp.date_of_joining.strftime("%Y-%m-%d") if emp.date_of_joining else ""
        writer.writerow([
            emp.employee_id or "",
            emp.name,
            emp.email,
            emp.role.value if hasattr(emp.role, "value") else str(emp.role),
            emp.department or "",
            emp.designation or "",
            emp.phone or "",
            doj_str,
            emp.status
        ])
        
    output.seek(0)
    
    filename = f"employees_export_{datetime.now().strftime('%Y%m%d%H%M%S')}.csv"
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/employees/bulk-deactivate")
async def bulk_deactivate_employees(
    payload: dict[str, list[str]],
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user)
):
    _require_hr_or_admin(user)
    ids = payload.get("user_ids", [])
    if not ids:
        raise ValidationError("No user_ids supplied in the request body.")
        
    result = await db.execute(
        select(User).where(User.id.in_(ids), User.tenant_id == user.tenant_id, User.deleted_at.is_(None))
    )
    targets = result.scalars().all()
    count = 0
    for target in targets:
        target.status = "Inactive"
        count += 1
        
    await db.commit()
    return {"message": f"Successfully deactivated {count} employees.", "deactivated_count": count}


@router.post("/employees/bulk-upload")
async def bulk_upload_employees(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user)
):
    _require_hr_or_admin(user)
    
    content = await file.read()
    filename = file.filename.lower()
    
    rows = []
    if filename.endswith(".csv"):
        text_content = content.decode("utf-8")
        csv_reader = csv.DictReader(io.StringIO(text_content))
        for row in csv_reader:
            rows.append(row)
    elif filename.endswith(".xlsx"):
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(cell).strip() if cell is not None else "" for cell in row]
            else:
                if any(cell is not None for cell in row):
                    row_dict = {}
                    for col_idx, cell in enumerate(row):
                        if col_idx < len(headers) and headers[col_idx]:
                            row_dict[headers[col_idx]] = cell
                    rows.append(row_dict)
    else:
        raise ValidationError("Unsupported file format. Please upload a .csv or .xlsx file.")

    added = 0
    updated = 0
    failed = 0
    errors = []

    # Map possible spreadsheet headers to system keys
    key_mappings = {
        "employee_id": ["employee id", "employee_id", "emp id", "empid", "id"],
        "name": ["name", "full name", "fullname", "employee name"],
        "email": ["email", "email address", "email_address"],
        "department": ["department", "dept"],
        "designation": ["designation", "designation/role", "title"],
        "phone": ["phone", "phone number", "phone_number", "contact"],
        "role": ["role", "user role", "user_role"],
        "date_of_joining": ["date of joining", "date_of_joining", "joining date", "doj"]
    }

    def get_mapped_value(row_dict: dict, canonical_key: str) -> str | None:
        aliases = key_mappings.get(canonical_key, [])
        for key in row_dict.keys():
            if str(key).strip().lower() in aliases:
                val = row_dict[key]
                return str(val).strip() if val is not None else None
        return None

    for idx, row in enumerate(rows):
        try:
            emp_id = get_mapped_value(row, "employee_id")
            name = get_mapped_value(row, "name")
            email = get_mapped_value(row, "email")
            dept = get_mapped_value(row, "department")
            desig = get_mapped_value(row, "designation")
            phone = get_mapped_value(row, "phone")
            role_str = get_mapped_value(row, "role") or "EMPLOYEE"
            doj_str = get_mapped_value(row, "date_of_joining")

            # Basic Validation
            if not email:
                raise ValueError("Email is required.")
            if not name:
                raise ValueError("Name is required.")
            if not emp_id:
                raise ValueError("Employee ID is required.")

            # Validate role
            try:
                role = UserRole(role_str.upper())
            except ValueError:
                role = UserRole.EMPLOYEE

            # Parse date of joining
            doj = None
            if doj_str:
                # Try standard formats
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
                    try:
                        # Strip time if parsed as string with time
                        cleaned_date = doj_str.split(" ")[0]
                        doj = datetime.strptime(cleaned_date, fmt).replace(tzinfo=timezone.utc)
                        break
                    except ValueError:
                        continue
                if not doj:
                    # Try datetime excel float or string
                    try:
                        # If openpyxl directly loaded it as datetime
                        if isinstance(row.get("Date of Joining"), datetime):
                            doj = row.get("Date of Joining").replace(tzinfo=timezone.utc)
                        elif isinstance(row.get("date_of_joining"), datetime):
                            doj = row.get("date_of_joining").replace(tzinfo=timezone.utc)
                    except Exception:
                        pass

            # Search if employee exists in tenant (by employee_id or email)
            result = await db.execute(
                select(User).where(
                    User.tenant_id == user.tenant_id,
                    User.deleted_at.is_(None),
                    or_(User.employee_id == emp_id, User.email == email)
                )
            )
            existing_user = result.scalar_one_or_none()

            # Verify email uniqueness globally to prevent conflicts across tenants
            global_res = await db.execute(
                select(User).where(
                    func.lower(User.email) == func.lower(email),
                    User.deleted_at.is_(None)
                )
            )
            global_user = global_res.scalar_one_or_none()
            if global_user and global_user.tenant_id != user.tenant_id:
                raise ValueError(f"Email '{email}' is already registered in another organization.")

            if existing_user:
                # If email matches another user but employee_id is different, block
                if existing_user.employee_id and existing_user.employee_id != emp_id:
                    raise ValueError(f"Email '{email}' matches another employee ID '{existing_user.employee_id}'.")
                
                # Update existing
                existing_user.name = name
                existing_user.email = email
                existing_user.employee_id = emp_id
                existing_user.department = dept
                existing_user.designation = desig
                existing_user.phone = phone
                existing_user.role = role
                if doj:
                    existing_user.date_of_joining = doj
                updated += 1
            else:
                # Create new
                # We need a default hashed password for login
                default_hashed = hash_password("Welcome@123")
                new_user = User(
                    tenant_id=user.tenant_id,
                    name=name,
                    email=email,
                    employee_id=emp_id,
                    department=dept,
                    designation=desig,
                    phone=phone,
                    role=role,
                    date_of_joining=doj,
                    hashed_password=default_hashed,
                    status="Active"
                )
                db.add(new_user)
                added += 1
        except Exception as e:
            failed += 1
            errors.append(f"Row {idx + 2}: {str(e)}")

    await db.commit()
    
    return {
        "success": True,
        "added": added,
        "updated": updated,
        "failed": failed,
        "errors": errors
    }


@router.get("/employees/{employee_uuid}", response_model=UserResponse)
async def get_employee_details(
    employee_uuid: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user)
):
    _require_hr_or_admin(user)
    result = await db.execute(
        select(User).where(User.id == employee_uuid, User.tenant_id == user.tenant_id, User.deleted_at.is_(None))
    )
    emp = result.scalar_one_or_none()
    if not emp:
        raise NotFoundError("Employee", employee_uuid)
    return UserResponse.model_validate(emp)


@router.put("/employees/{employee_uuid}", response_model=UserResponse)
async def update_employee(
    employee_uuid: str,
    payload: UserUpdate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user)
):
    _require_hr_or_admin(user)
    return await UserService.update_user(db, user.tenant_id, employee_uuid, payload)


@router.patch("/employees/{employee_uuid}/deactivate", response_model=UserResponse)
async def toggle_employee_status(
    employee_uuid: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user)
):
    _require_hr_or_admin(user)
    result = await db.execute(
        select(User).where(User.id == employee_uuid, User.tenant_id == user.tenant_id, User.deleted_at.is_(None))
    )
    emp = result.scalar_one_or_none()
    if not emp:
        raise NotFoundError("Employee", employee_uuid)
    
    # Toggle status
    emp.status = "Inactive" if emp.status == "Active" else "Active"
    await db.commit()
    await db.refresh(emp)
    return UserResponse.model_validate(emp)


# ─── Employee Profile Endpoints (Any Logged-in User) ──────────────────────────

@router.get("/profile", response_model=UserResponse)
async def get_own_profile(user: User = Depends(get_current_user)):
    return UserResponse.model_validate(user)


@router.post("/profile-update-request", response_model=ProfileUpdateRequestResponse)
async def request_profile_update(
    payload: ProfileUpdateRequestCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user)
):
    # Retrieve current value dynamically
    old_val = getattr(user, payload.field, None) if hasattr(user, payload.field) else None
    if isinstance(old_val, datetime):
        old_val = old_val.strftime("%Y-%m-%d")
    elif old_val is not None:
        old_val = str(old_val)

    # Restrict what fields can be requested for updates to avoid messing up critical database states
    allowed_fields = ["phone", "email", "name", "designation", "department", "profile_photo"]
    if payload.field not in allowed_fields:
        raise ValidationError(f"Updates for field '{payload.field}' are not permitted.")

    if payload.field == "email":
        # Check email uniqueness globally (case-insensitive)
        email_res = await db.execute(
            select(User).where(
                func.lower(User.email) == payload.new_value.lower(),
                User.id != user.id,
                User.deleted_at.is_(None)
            )
        )
        if email_res.scalar_one_or_none():
            raise ConflictError(f"User with email '{payload.new_value}' already exists.")

    req = ProfileUpdateRequest(
        tenant_id=user.tenant_id,
        user_id=user.id,
        field=payload.field,
        old_value=old_val,
        new_value=payload.new_value,
        reason=payload.reason,
        status="Pending"
    )
    db.add(req)
    await db.commit()
    await db.refresh(req)
    
    # Attach user model to response object for validation
    req.user = user
    return ProfileUpdateRequestResponse.model_validate(req)


@router.get("/profile-update-requests/my", response_model=list[ProfileUpdateRequestResponse])
async def list_own_profile_requests(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user)
):
    result = await db.execute(
        select(ProfileUpdateRequest)
        .where(ProfileUpdateRequest.user_id == user.id)
        .order_by(ProfileUpdateRequest.created_at.desc())
    )
    requests = result.scalars().all()
    for req in requests:
        req.user = user
    return [ProfileUpdateRequestResponse.model_validate(r) for r in requests]


# ─── Profile Update Request Actions (Admin/HR Only) ──────────────────────────

@router.get("/profile-update-requests", response_model=list[ProfileUpdateRequestResponse])
async def list_all_profile_requests(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    status: str | None = Query(None)
):
    _require_hr_or_admin(user)
    
    q = select(ProfileUpdateRequest).where(ProfileUpdateRequest.tenant_id == user.tenant_id)
    if status:
        q = q.where(ProfileUpdateRequest.status == status)
    
    q = q.order_by(ProfileUpdateRequest.created_at.desc())
    
    result = await db.execute(q)
    requests = result.scalars().all()
    
    # We load matching users to prevent validation errors in user field
    response_list = []
    for req in requests:
        user_res = await db.execute(select(User).where(User.id == req.user_id))
        req.user = user_res.scalar_one_or_none()
        response_list.append(req)
        
    return [ProfileUpdateRequestResponse.model_validate(r) for r in response_list]


@router.post("/profile-update-requests/{request_id}/review", response_model=ProfileUpdateRequestResponse)
async def review_profile_request(
    request_id: str,
    payload: ProfileUpdateRequestReview,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user)
):
    _require_hr_or_admin(user)
    
    result = await db.execute(
        select(ProfileUpdateRequest).where(
            ProfileUpdateRequest.id == request_id, 
            ProfileUpdateRequest.tenant_id == user.tenant_id
        )
    )
    req = result.scalar_one_or_none()
    if not req:
        raise NotFoundError("ProfileUpdateRequest", request_id)
        
    if req.status != "Pending":
        raise ConflictError(f"Request is already reviewed and marked as '{req.status}'.")
        
    if payload.status not in ["Approved", "Rejected"]:
        raise ValidationError("Status must be either 'Approved' or 'Rejected'.")
        
    req.status = payload.status
    req.reviewed_by = user.id
    req.review_notes = payload.review_notes
    req.updated_at = datetime.now(timezone.utc)
    
    user_res = await db.execute(select(User).where(User.id == req.user_id))
    target_user = user_res.scalar_one_or_none()
    if not target_user:
        raise NotFoundError("User target for correction not found.")
        
    if payload.status == "Approved":
        # Dynamic update on User model!
        if hasattr(target_user, req.field):
            if req.field == "email":
                # Check email uniqueness globally (case-insensitive) to prevent cross-tenant collisions
                email_res = await db.execute(
                    select(User).where(
                        func.lower(User.email) == req.new_value.lower(),
                        User.id != target_user.id,
                        User.deleted_at.is_(None)
                    )
                )
                if email_res.scalar_one_or_none():
                    raise ConflictError(f"User with email '{req.new_value}' already exists.")

            if req.field == "date_of_joining":
                # Parse date
                try:
                    dt = datetime.strptime(req.new_value, "%Y-%m-%d").replace(tzinfo=timezone.utc)
                    setattr(target_user, req.field, dt)
                except ValueError:
                    raise ValidationError(f"Invalid date format for '{req.field}': must be YYYY-MM-DD.")
            else:
                setattr(target_user, req.field, req.new_value)
        else:
            raise ValidationError(f"Field '{req.field}' does not exist on user database record.")
            
    await db.commit()
    await db.refresh(req)
    req.user = target_user
    return ProfileUpdateRequestResponse.model_validate(req)
